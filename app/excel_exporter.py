import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ExcelExporter:
    """Classe para exportar dados para o Excel."""

    def __init__(self, data):
        self.data = data

    def export_to_excel(self, output):
        """Exporta os dados para um arquivo Excel."""
        try:
            workbook = Workbook()
            sheet = workbook.active

            # Escrever os cabeçalhos na primeira linha
            headers = ['Data', 'Localizador', 'Origem', 'Destino', 'Passageiros', 'Milhas', 'Taxas', 'Tipo de movimentação']
            sheet.append(headers)

            # Aplicar formatação aos cabeçalhos
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            # Escrever os dados nas linhas subsequentes
            for item in self.data:
                row = [
                    item.get('Data', ''),
                    item['Localizador'],
                    item['Origem'],
                    item['Destino'],
                    item['Passageiros'],
                    item['Milhas'],
                    item['Taxas'],
                    item['Tipo de movimentação']
                ]
                sheet.append(row)

            # Aplicar formatação condicional às células
            for row in sheet.iter_rows(min_row=2, min_col=6, max_col=7):
                for cell in row:
                    if cell.value:
                        cell.number_format = '#,##0.00'

            # Ajustar a largura das colunas automaticamente
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(output)
        except Exception as error:
            logging.error(f"Erro ao exportar dados para o Excel: {error}")
            raise